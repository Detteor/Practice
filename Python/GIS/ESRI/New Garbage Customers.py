import shutil
import openpyxl
from openpyxl import load_workbook
import arcpy
import os
from arcpy import env
from arcpy import da
from datetime import date
import logging
from os.path import join as pj
import sys
import smtplib, os.path as op, base64, mimetypes
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import  formatdate
from email import encoders
from datetime import datetime as dt
from datetime import timedelta
import getpass

from dotenv import find_dotenv, load_dotenv

load_dotenv(find_dotenv())
gis_user = "GIS_USER"
gis_password = "GIS_PASSWORD"
work_email = "WORK_EMAIL"
email_server = "EMAIL_SERVER"
work_email_2 = "WORK_EMAIL_2"
database_user = "DB_USER_NAME"

# ----------------------------------------------------------------------------------- #

# Varibles to update if the script doesn't work!
# - 1) Update this to correct source excel as needed -
sourcePath = 'path to source'
sourceFile = 'source file needed'

# - 2) Update this to correct destination path if the .py folder is moved -
destinationPath = path = 'path to script'

# - 3) Update this whenever the 'month' changes in the source sheet (not always accurate to current month) -
sourceSheetName = u'April 2021'

# ----------------------------------------------------------------------------------- #

# Normal Variables
# - This is the saved file name of the excel workbook that gets editted -
workingCopyName = 'path to edited workbook'

noNewCustomers = False

# Database Connections
engSDE = "C:\\Users\\" + getpass.getuser() + "\\AppData\\Roaming\\ESRI\\Desktop10.5\\ArcCatalog\\Engineering.sde"
prodSDE = "C:\\Users\\" + getpass.getuser() + "\\AppData\\Roaming\\ESRI\\Desktop10.5\\ArcCatalog\\PROD.sde"

# Swap to False when done testing
testing = False

# ----------------------------------------------------------------------------------- #

# Functions
# Creates a list with Trash District Info
def createTrashDefList(path, file):
    logs.info('createTrashDefList - Creating Trash District Info List')
    wb = load_workbook(path + file)

    # - This should match what the 'correct' sheet is in 'TRASH - DISTRICT MASTER LIST.xlsx', it shouldn't regularly change. -
    ws = wb.get_sheet_by_name(u'Districts')

    dim = ws.max_row

    districtInfo = []
    for i in range(2, dim, 1):
        if len(str(ws.cell(i, 3).value)) == 1:
            distID = 'T00' + str(ws.cell(i, 3).value)
        if len(str(ws.cell(i, 3).value)) == 2:
            distID = 'T0' + str(ws.cell(i, 3).value)
        if len(str(ws.cell(i, 3).value)) == 3:
            distID = 'T' + str(ws.cell(i, 3).value)
        if ws.cell(i, 6).value == 1:
            distZone = 'REPULBIC'
        if ws.cell(i, 6).value == 2:
            distZone = 'WASTE INDUSTRIES'
        distDayCap = ws.cell(i, 7).value
        distDayCap = distDayCap.capitalize()
        if ws.cell(i, 8).value == 'Yard Waste Service':
            distYW = True
        if ws.cell(i, 8).value == 'NO Yard Waste Service':
            distYW = False
        districtInfo.append([distID, distDayCap, distZone, distYW, ws.cell(i, 4).value])
    logs.info('createTrashDefList - Creating Trash District Info List - Complete')
    return districtInfo

# Create Logger
# Source Code: 'Q:\Workspace\PythonManuals\PWReports_Local.py'
def initLog(log_file=None):
    log = logging.getLogger("New Garbage Customers Log")
    log.setLevel(logging.INFO)

    h1 = logging.FileHandler(log_file)
    h2 = logging.StreamHandler()

    f = logging.Formatter("[%(levelname)s] [%(asctime)s] [%(lineno)d] - %(message)s",'%m/%d/%Y %I:%M:%S %p')

    h1.setFormatter(f)
    h2.setFormatter(f)

    h1.setLevel(logging.INFO)
    h2.setLevel(logging.INFO)

    log.addHandler(h1)
    log.addHandler(h2)
    
    return log

# ----------------------------------------------------------------------------------- #

# Main
# Start Logger
logs = initLog(path + "New Garbage Customers Log.txt")
logs.info('# ----------------------------------------------------------------------------------- #')
logs.info("Creating Log at {}".format(path))
logs.info('# ----------------------------------------------------------------------------------- #')

# Loading Excel File (creating a copy)
logs.info('Copying Source File: {}'.format(sourceFile))
shutil.copy(sourcePath + sourceFile, destinationPath + 'NGC_workingCopy.xlsx')
logs.info('Copying Source File: {} - Complete'.format(sourceFile))

wb = load_workbook(path + workingCopyName)

ws = wb.get_sheet_by_name(sourceSheetName)
dim = ws.max_row

# Deleting rows that are either not active, or already verified
logs.info('Removing non-new / non-verified rows')
for i in range(dim, 1, -1):
    vendor = ws.cell(i, 19).value
    verified = ws.cell(i, 23).value
    if(vendor != 'W' and vendor != 'R'):
        ws.delete_rows(i)
    if(vendor == 'W' and verified == 'YES'):
        ws.delete_rows(i)
    if(vendor == 'R' and verified == 'YES'):
        ws.delete_rows(i)
        
if ws.max_row == 1:
    noNewCustomers = True

ws.delete_rows(1)
logs.info('Removing non-new / non-verified rows - Complete')
newDim = ws.max_row

wb.save(path + 'NGC_workingCopy.xlsx')

logs.info('Creating newCust list')
newCust = []
for i in range(newDim, 0, -1):
    if i == 0:
        break
    #ParcelID, Address, Zip, City, State, Dist No., Vendor
    newCust.append([ws['C' + str(i)].value, ws['L' + str(i)].value, ws['O' + str(i)].value, ws['M' + str(i)].value, ws['N' + str(i)].value, ws['E' + str(i)].value, ws['S' + str(i)].value])
logs.info('Creating newCust list - Complete')

env.overwriteOutput = True
env.qualifiedFieldNames = True

A = prodSDE + "\\PROD.GISADMIN.TaxParcels"
idFieldA = 'Name'
idFieldLot = 'Lot_Number'
B = engSDE + '\\Engineering.DBO.Garbage\\Engineering.DBO.Customers'
idFieldB = 'PARCELID'

logs.info('Creating Tax Parcel Geometry Dictionary') 
geometries = {key:value for (key,value) in arcpy.da.SearchCursor(A, [idFieldA, 'SHAPE@'])}
logs.info('Creating Tax Parcel Geometry Dictionary - Complete')
logs.info('Creating Tax Parcel Lot Dictionary - Complete')
lots = {key:value for (key,value) in arcpy.da.SearchCursor(A, [idFieldA, idFieldLot])}
logs.info('Creating Tax Parcel Lot Dictionary - Complete')
logs.info('Creating Customer Index / Assigning newCustomer Index Cursor') 
customers = [x for x in arcpy.da.SearchCursor(B, idFieldB)]
newCustomer = arcpy.da.InsertCursor(B, ['PARCELID', 'Block', 'AddressNo', 'Address', 'Zip', 'City', 'State', 'GDistID', 'GZone', 'G', 'Garbage3', 'Recycle', 'Recycle_2',
                                        'YardWaste', 'Yard_Waste_2', 'Symbology', 'GNotes', 'ActStatus', 'LotStatus', 'GTaxStatus', 'GTaxUnits',
                                        'GTaxDate', 'Multiple_Units', 'DISABLED_SRVC', 'AddressPre', 'AddressRoad', 'AddressSuf',
                                        'GCollection', 'YW', 'GDistName', 'SHAPE@'])
logs.info('Creating Customer Index / Assigning newCustomer Index Cursor - Complete') 

edit = da.Editor(engSDE)
edit.startEditing(True, True)
edit.startOperation()

file = 'TRASH - DISTRICT MASTER LIST.xlsx'
districtList = createTrashDefList(path, file)

logs.info('Inserting New Customers')
newCounter = []
newCounterLots = []
existingCounter = []
existingCounterLots = []
if noNewCustomers == True:
    logs.error('No new or "waiting to be verified" customers!')
if noNewCustomers == False:
    for x in newCust:
        existing = 0
        for row in customers:
            if x[0] == row[0][:18]:
                logs.info('Existing!: ' + str(row[0][:18]))
                existing = 1
                existPARCELID = row[0][:18] + str('-00001')
                existingCounter.append(x[0])
                existingCounterLots.append(lots[existPARCELID])
        if existing == 0:
            newCounter.append(str(x[0]))
            logs.info('New!: ' + str(x[0]))
            PARCELID = x[0] + str('-00001')
            Block = x[0][0:13]
            Address = x[1]
            numList = [int(i) for i in Address.split() if i.isdigit()]
            AddressNo = numList[0]
            Zip = x[2]
            City = x[3]
            State = x[4]
            GDistID = x[5][:4]
            if x[6] == 'R': GZone = 'REPUBLIC'
            if x[6] == 'W': GZone = 'WASTE INDUSTRIES'
            Garb = Rec = None
            Garb2 = Rec2 = 'NONE'
            GNotes = 'NC'
            ActStatus = 'ACTIVE'
            LotStatus = 'Dwelling Unit'
            GTaxStatus = 'Taxed'
            GTaxUnits = 1
            Today = date.today()
            GTaxDate = Today.strftime("%d/%m/%Y")
            Multiple_Units = '0'
            DISABLED_SRVC = '0'
            
            wordList = Address.split()
            wordList.pop(0)
            preList = ['NORTH', 'SOUTH', 'EAST', 'WEST']
            if len(wordList) == 2:
                AddressSuf = wordList[1]
                AddressRoad = wordList[0]
                AddressPre = 'None'
            else:
                if wordList[0] in preList:
                    AddressPre = wordList[0]
                    road = wordList[1:-1]
                else:
                    AddressPre = 'None'
                    road = wordList[:-1]
                AddressSuf = wordList[-1]
                AddressRoad = ' '.join([str(elem) for elem in road])
                        
            for row in districtList:
                if GDistID == row[0]:
                    GCollection = row[1]
                    if row[3] == True:
                        YW = 'Yes'
                        Symbology = 'G R Y'
                        Yard = None
                        Yard2 = 'NONE'
                    if row[3] == False:
                        YW = 'No'
                        Symbology = 'G R'
                        Yard = Yard2 = 'NOT AVALIBLE'
                    GDistName = row[4]
                    
            shape = geometries[PARCELID]
            newCounterLots.append(str(lots[PARCELID]))
            newCustomer.insertRow((PARCELID, Block, AddressNo, Address, Zip, City, State, GDistID, GZone, Garb, Garb2, Rec, Rec2,
                                   Yard, Yard2, Symbology, GNotes, ActStatus, LotStatus, GTaxStatus, GTaxUnits,
                                   GTaxDate, Multiple_Units, DISABLED_SRVC, AddressPre, AddressRoad, AddressSuf,
                                   GCollection, YW, GDistName, shape))
logs.info('Inserting New Customers - Complete') 

logs.info("Update Finished. Reconciling and Posting.")

arcpy.ReconcileVersions_management(engSDE, 
                                 "ALL_VERSIONS", 
                                 "dbo.DEFAULT", 
                                 database_user, 
                                 "LOCK_ACQUIRED",
                                 "NO_ABORT",
                                 "BY_OBJECT",
                                 "FAVOR_TARGET_VERSION",
                                 "POST",
                                 "KEEP_VERSION",
                                 r"c:\Customer_Geo_Update.txt")

logs.info("Posting Complete. Sending Email.")

# These are for testing purposes.
if testing:
    newCounter = ['0-00-00000-00-0000', '0-00-00000-00-0000', '0-00-00000-00-0000', '0-00-00000-00-0000', '0-00-00000-00-0000']
    newCounterLots = ['11', '24', '124', '223', '14']
    existingCounter = ['1-11-11111-11-1111', '1-11-11111-11-1111', '1-11-11111-11-1111']
    existingCounterLots = ['1', '4', '24']

# Send notification email if a new customer was entered and a different one if not.
# Code Below from PWReports_Local.py 
#Process: Report To Email
server = email_server
sent_from = work_email
if not testing:
    sent_to = [work_email, work_email_2]
if testing:
    sent_to = [work_email]
today = dt.now()
subject = "New Garbage Customers: " + today.strftime("%B %d, %Y %I:%M:%S %p")
if len(newCounter) > 0 and len(existingCounter) == 0:
    intro = "Hello! \r\n\nHere is a list of all new customers that were created in the system today (Customers.DBO):"
    body = '\r\n\n'
    for i in range(0, len(newCounter), 1):
        body = body + str(newCounter[i]) + '   ' + str(newCounterLots[i]) + "\r\n"
    body = body + '\nThere are no "waiting to be verified" customers.\r\n'
    closing = "\nThank You, \r\n{}".format(gis_user)
    text = intro + body + closing

if len(newCounter) == 0 and len(existingCounter) > 0:
    intro = "Hello! \r\n\nHere is a list of all new customers that were created in the system today (Customers.DBO):"
    body = '\r\n\nThere are no new customers at this time (' + today.strftime("%B %d, %Y %I:%M:%S %p") + ').\r\n\nThere are customers that are "waiting to be verified":\r\n\n'
    for i in range(0, len(existingCounter), 1):
        body = body + str(existingCounter[i]) + '   ' + str(existingCounterLots[i]) + "\r\n"
    closing = "\nThank You, \r\n{}".format(gis_user)
    text = intro + body + closing

if len(newCounter) > 0 and len(existingCounter) > 0:
    intro = "Hello! \r\n\nHere is a list of all new customers that were created in the system today (Customers.DBO):"
    body = '\r\n\n'
    for i in range(0, len(newCounter), 1):
        body = body + str(newCounter[i]) + '   ' + str(newCounterLots[i]) + "\r\n"
    body = body + '\nThere are also customers that are "waiting to be verified":\r\n\n'
    for i in range(0, len(existingCounter), 1):
        body = body + str(existingCounter[i]) + '   ' + str(existingCounterLots[i]) + "\r\n"
    closing = "\nThank You, \r\n{}".format(gis_user)
    text = intro + body + closing

if len(newCounter) == 0 and len(existingCounter) == 0:
    intro = "Hello! \r\n\nHere is a list of all new customers that were created in the system today (Customers.DBO):"
    body = '\r\n\nThere are no new or "waiting to be verified" customers at this time (' + today.strftime("%B %d, %Y %I:%M:%S %p") + ').\r\n'
    closing = "\nThank You, \r\n{}".format(gis_user)
    text = intro + body + closing

## Use MIMEMultipart to build email
msg = MIMEMultipart()
msg['From'] = sent_from
msg['To'] = ", ".join(sent_to)
msg['Date'] = formatdate(localtime = True)
msg['Subject'] = subject

## Add Message Text
message = MIMEBase('text', 'plain')
message.set_payload(text)
encoders.encode_base64(message)
msg.attach(message)

server = smtplib.SMTP(server)
server.ehlo()
##server.starttls()
server.sendmail(sent_from, sent_to, msg.as_string())
server.quit()

edit.stopOperation()
edit.stopEditing(True)