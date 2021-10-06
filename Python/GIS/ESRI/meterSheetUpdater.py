from arcgis import GIS
from datetime import datetime
from openpyxl import load_workbook
from os import remove
from pathlib import Path
import warnings
from time import time
import requests
from dotenv import find_dotenv, load_dotenv

load_dotenv(find_dotenv())
gis_user = "GIS_USER"
gis_password = "GIS_PASSWORD"


warnings.simplefilter("ignore")

#Pulls water levels for the day in question, then averages them out
def waterLevel(date):
    payload = {'format':'json',
                'site':'385041075395602',
                'startDt':date,
                'endDT':date}
    r = requests.get('https://waterservices.usgs.gov/nwis/iv/', params=payload).json()['value']['timeSeries'][0]['values'][0]['value']
    avgLevel = 0
    for item in r:
        avgLevel += float(item['value'])
    avgLevel = avgLevel / len(r)
    avgLevel = format(avgLevel, '.2f')
    avgLevel = float(avgLevel)
    readingDate = r[0]['dateTime']
    print(f'The water table level is {avgLevel}ft below land and was read at {readingDate}')
    return [avgLevel, readingDate]

#Pull rain data across each of the county stations, then averages for the region
def CoCoRaHs(date, station, weekLength=7):
    payload = {'Format':'JSON',
               'State':'DE',
               'StartDate':str(date - datetime.timedelta(days=weekLength)),
               'EndDate':str(date),
               'Station':str(station)}
    r = requests.get('https://data.cocorahs.org/export/exportreports.aspx', params=payload).json()['data']
    totalPercip = 0
    if len(r['reports']) <= 5 + (weekLength-7):
        return None
    for row in r['reports']:
        if row['totalpcpn'] == -1:
            continue
        totalPercip += row['totalpcpn']
    totalPercip = format(totalPercip, '.2f')
    totalPercip = float(totalPercip)
    return(totalPercip)

#Pulls Survey123 data as excel files (saving to exportLoc)
def grabSurvey123(meterDate, direction, exportLoc):
    print('Beginning ' + direction + ' Route Survey123 Download.')
    print('Logging into .')
    agoLogin = GIS(url=None,
                          username=gis_user,
                          password=gis_password)

    month = meterDate[:2]
    day = meterDate[3:5]
    year = meterDate[6:10]

    featureID = ""
    if direction == "North":
        featureID = "6e4bcc172e844736ad1aa8505537aed4"
    if direction == "South":
        featureID = "22af1b6e3ff04c0497fc12b57692b0e8"

    itemToDownload = agoLogin.content.get(featureID)
    itemExportName = month + day
    exportParameters = {"layers": [ {"id" : 0, "where" : "CreationDate > '"+ meterDate + " 12:00:00 AM' AND CreationDate < '"+ meterDate + " 11:59:59 PM'"} ] }

    print('Generating Excel in ArcGIS Online.')
    itemToDownload.export(title=itemExportName,
                          export_format="Excel",
                          parameters=exportParameters,
                          wait=True)

    searchForExportedItem = agoLogin.content.search(query=itemExportName)
    exportedItemID = searchForExportedItem[0].id
    getTheExportedItem = agoLogin.content.get(exportedItemID)

    print('Downloading to: ' + exportLoc)
    getTheExportedItem.download(save_path=exportLoc,
                                file_name="{}.xlsx".format(itemExportName))

    print('Download Complete! Removing ArcGIS Online file.')
    getTheExportedItem.delete()

#Downloads pictures from Survey123
def downloadPictures(meterDate, direction, pictureLoc):
    print('Beginning ' + direction + ' Route Survey123 Image Download.')
    print('Logging into ArcGIS.')
    agoLogin = GIS(url=None,
                          username=gis_user,
                          password=gis_password)
    
    featureID = ""
    if direction == "North":
        featureID = "6e4bcc172e844736ad1aa8505537aed4"
    if direction == "South":
        featureID = "22af1b6e3ff04c0497fc12b57692b0e8"

    myFLItem = agoLogin.content.get(featureID)
    attachmentsLayer = myFLItem.layers[0]
    oidField = attachmentsLayer.properties.objectIdField
    myRecords = attachmentsLayer.query(where="CreationDate > '"+ meterDate + " 12:00:00 AM' AND CreationDate < '"+ meterDate + " 11:59:59 PM'",out_fields=oidField)

    for r in myRecords.features:
        myOID = r.get_value(oidField)
        attachments = attachmentsLayer.attachments.get_list(myOID)
        for attachment in attachments:
            myDownload = attachmentsLayer.attachments.download(oid=myOID, attachment_id=attachment['id'],save_path=pictureLoc)
            print ("Downloaded:" + myDownload[0])

#Inserts info into the flow master sheet
def insertIntoFlowSheet(wbInsert, wbMaster, wbMasterDO, meterDate, savepath, rainPrompt, averageRain, waterPrompt, waterInfo, test=False):
    rowDic = {
        "PS35" : "*ps35_east_dover",
        "COD-5" : "cod_5_white_oak",
        "COD-7" : "cod_7_dover_downs",
        "PS23b" : "ps23b_kentwood",
        "PS23" : "*ps23_kentwood",
        "PS30" : "*ps30_northeast_dykes_b_",
        "PS16" : "*ps_16_dykes_branch_hatchery_",
        "PS20D" : "*ps20d_spring_meadows",
        "LS28" : "ps28_willis_auto_mall",
        "PS28a" : "ps28a_burtonwood_villiage",
        "PS1" : "ps1_water_*sewer",
        "PS24B" : "*ps24b_wicksfield",
        "PS24a" : "ps24a_smyrnatownecenter",
        "PS24" : "*ps24_holly_hills",
        "PS10" : "*ps10_garrison's_lake",
        "PS10b" : "*ps10b_garrison's_lake",
        "PS12" : "*ps12_cheswold",
        "PS32" : "*ps32_nobles_pond",
        "PS2" : "ps2_main_transmission",
        "PS19" : "*ps19_carlisle_village",
        "PS19b" : "*ps19b_forty_nine_pines",
        "PS27 COD33" : "cod33_puncheon_run",
        "M14b" : "PS14b_CWSWA_OLD_TRTMNT_",
        "LS14d" : "PS14d_CWSWA_NELLIE_STOKES",
        "PS36" : "PS36_DSWA",
        "M14a" : "*m_14a_cwswa_+rv_right_",
        "PS14" : "*ps_14_isaacs_branch_left_",
        "M14c" : "m14c_rodney_vlg_cod34_",
        "PS1a" : "",
        "TOS-1" : "*tos_1_gateway_north",
        "LS12a" : "ls12a_1886_dover_llc_",
        "PS23a" : "",
        "PS3a" : "",
        "PS3" : "*ps3_dover_main",
        "PS29" : "*ps29_n_little_creek",
        "COD-2" : "",
        "PS22" : "*ps22_capitol_park",
        "PS5" : "*ps_5_general's_way",
        "COD-32" : "",
        "PS6" : "ps_6_dafb",
        "WLU" : "",
        "PS15" : "ps15_tidbury",
        "LS15a" : "ps15a_mifflin_meadows",
        "PS9A" : "ps9a_riverside",
        "PS9" : "ps9_eagle_meadows",
        "PS4" : "",
        "PS44" : "*PS44_BARRETT_FARMS",
        "PS41" : "ps41_longacre_village",
        "PS40" : "ps40_fox_hollow",
        "PS42" : "ps42_olde_field_village",
        "PS45" : "ps_45_riverview",
        "PS11" : "*PS11_MAGNOLIA",
        "PS8" : "*PS8_LITTLE_HAVEN",
        "PS13" : "",
        "PS13a" : "PS13a_ILC",
        "PS21" : "*PS21_FELTON",
        "LS21b" : "PS21b_LAKE_FOREST",
        "PS26" : "PS26_COLONY_SOUTH",
        "PS7" : "*PS7_MILFORD",
        "M7a" : "*PS7a_MILFORD_METER",
        "PS7b" : "",
        "PS34" : "*PS34_SOUTHFIELD",
        "PS17" : "*PS17_HARRINGTON",
        "PS31A" : "PS31a_SPORTS_COMPLEX",
        "PS31" : "*PS31_WATERS_EDGE",
        "M36a" : "DSWA_RECOVERY_METER",
        "PS37" : "*PS37_MILFORD_NECK",
        "PS12f" : "",
        "PS20c" : "*ps20c_brenford_station",
        "PS20b" : "*ps20b_garrisons_lake_green",
        "PS5d" : "ps_5d_microtel"
        }
    
    month = meterDate[:2]
    day = meterDate[3:5]
    year = meterDate[6:10]
    formattedDate = year + '-' + month + '-' + day + ' 00:00:00'
    
    wsMaster = wbMaster.get_sheet_by_name(u'Datasheet')
    wbMasterDO = wbMasterDO.get_sheet_by_name(u'Datasheet')
    wsInsert = wbInsert.active
    workingCol = ""
    for col in wbMasterDO.iter_cols(1, None, 5, 5):
        for cell in col:
            if str(cell.value) == formattedDate:
                workingCol = cell.column
    if workingCol == "":
        msg = "ERROR: Date not found in master sheet. Were meters read on the normal day?"
        raise Exception(msg)

    dimMaster = wsMaster.max_row
    dimInsert = wsInsert.max_row

    #Inserting rain data
    if rainPrompt == 'Y' or rainPrompt == 'y':
        wsMaster.cell(86, workingCol).value = str(averageRain) + '"Rain'

    #Inserting water data
    if waterPrompt == 'Y' or waterPrompt == 'y':
        wsMaster.cell(87, workingCol).value = str(waterInfo[0])
    
    for i in range(8, dimMaster+1):
        if wsMaster.cell(i, 6).value == None:
            continue
        for j in range(2, dimInsert+1):
            if rowDic[str(wsMaster.cell(i, 6).value)] == str(wsInsert.cell(j, 8).value):
                print("MATCH!: " + rowDic[str(wsMaster.cell(i, 6).value)])
                if rowDic[str(wsMaster.cell(i, 6).value)] == "ps40_fox_hollow":
                    wsMaster.cell(i, workingCol).value = float(wsInsert.cell(j, 9).value)*1000
                elif rowDic[str(wsMaster.cell(i, 6).value)] == "*ps20c_brenford_station":
                    wsMaster.cell(i, workingCol).value = float(wsInsert.cell(j, 9).value)*1000
                elif rowDic[str(wsMaster.cell(i, 6).value)] == "ps41_longacre_village":
                    wsMaster.cell(i, workingCol).value = float(wsInsert.cell(j, 9).value)*1000
                elif rowDic[str(wsMaster.cell(i, 6).value)] == "*tos_1_gateway_north":
                    wsMaster.cell(i, workingCol).value = float(wsInsert.cell(j, 9).value)+10000000
                elif rowDic[str(wsMaster.cell(i, 6).value)] == "PS14d_CWSWA_NELLIE_STOKES":
                    wsMaster.cell(i, workingCol).value = float(wsInsert.cell(j, 9).value)+100000000
                else:
                    wsMaster.cell(i, workingCol).value = float(wsInsert.cell(j, 9).value)
                    
    if test:
        wbMaster.save(savepath + "\Meter_Flow_2021_TEST_COPY.xlsx")
    else:
        wbMaster.save(savepath + "\Meter_Flow_2021.xlsx")


#Main ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
meterDate = input("Enter the date of the Meter Run (MM/DD/YYYY): ")
month = meterDate[:2]
day = meterDate[3:5]
year = meterDate[6:10]
testLoc = r"path to script"
mainLoc = r"path to main location"
exportLocNorth = r"path to north export"
exportLocSouth = r"path to south export"
pictureLoc = r"path to meter photo location" + str(month) + str(day)
if not os.path.exists(pictureLoc):
    os.makedirs(pictureLoc)

if int(month) not in list(range(1,13)):
    raise Exception("Invalid Month!")

if int(day) not in list(range(1,32)):
    raise Exception("Invalid Day!")

try: int(year)
except: raise Exception("Invalid Year!")

# Water Level ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

waterPrompt = input('Update depth to water table measurement? (Y/N): ')
if waterPrompt == 'Y' or waterPrompt == 'y':
    waterDate = year + '-' + month + '-' + day
    waterInfo = waterLevel(waterDate)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# Rain ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
rainPrompt = input('Pull rain data from CoCoRaHs? (Y/N): ')

if rainPrompt == 'Y' or rainPrompt == 'y':
    weekLengthPrompt = input('Is this a 7 day reading? (Y/N): ')
    if weekLengthPrompt == 'Y' or weekLengthPrompt == 'y':
        weekLength = 7
    else:
        weekLength = input('How many days ago was the last reading?: ')
    dateObj = datetime.strptime(meterDate, '%m/%d/%Y')
    perc = []
    stationList = ['DE-KN-03', 'DE-KN-24', 'DE-KN-28', 'DE-KN-12', 'DE-KN-31', 'DE-KN-32', 'DE-SS-30', 'DE-KN-13', 'DE-KN-20']
    #DE-KN-03 Smyrna
    #DE-KN-24 Felton
    #DE-KN-28 Magnolia
    #DE-KN-12 Camden
    #DE-KN-31 Dover
    #DE-KN-32 Dover
    #DE-SS-30 Milford
    #DE-KN-13 Milford
    #DE-KN-20 Clayton
    for item in stationList:
        rain = CoCoRaHs(dateObj, item, weekLength)
        if rain == None:
            continue
        else:
            perc.append(rain)

    averageRain = round(sum(perc) / len(perc), 3)
    print('Average Rain: ' + str(averageRain) + 'in.')
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


#Survey123 ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
northPrompt = input("Pull North Route Survey123 excel? (Y/N): ")
southPrompt = input("Pull South Route Survey123 excel? (Y/N): ")

startTime = time()

if northPrompt == 'Y' or northPrompt == 'y':
    pathN = Path(f"{exportLocNorth}\\{month}{day}.xlsx")
    if pathN.exists:
        print("Exists!")
        remove(pathN)
    grabSurvey123(meterDate, 'North', exportLocNorth)
if southPrompt == 'Y' or southPrompt == 'y':
    pathS = Path(f"{exportLocSouth}\\{month}{day}.xlsx")
    if pathS.exists:
        print("Exists!")
        remove(pathS)
    grabSurvey123(meterDate, 'South', exportLocSouth)

testPrompt = input("Testing? (Y/N): ")
if testPrompt == 'Y' or testPrompt == 'y':
    test = True
    savepath = testLoc
    wbMaster = load_workbook(testLoc + "\Meter_Flow_2021_TEST_COPY.xlsx")
    wbMasterDO = load_workbook(testLoc + "\Meter_Flow_2021_TEST_COPY.xlsx", data_only = True)
elif testPrompt == 'N' or testPrompt == 'n':
    test = False
    savepath = mainLoc
    wbMaster = load_workbook(mainLoc + "\Meter_Flow_2021.xlsx")
    wbMasterDO = load_workbook(mainLoc + "\Meter_Flow_2021.xlsx", data_only = True)
else:
    raise Exception("Please input (Y/N) for the testing prompt.")
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

#Loading Workbooks and inserting into sheet ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
northWB = load_workbook(exportLocNorth + "\\" + month + day + ".xlsx")
southWB = load_workbook(exportLocSouth + "\\" + month + day + ".xlsx")

insertIntoFlowSheet(northWB, wbMaster, wbMasterDO, meterDate, savepath, rainPrompt, averageRain, waterPrompt, waterInfo, test)
insertIntoFlowSheet(southWB, wbMaster, wbMasterDO, meterDate, savepath, rainPrompt, averageRain, waterPrompt, waterInfo, test)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

#Downloading Pictures ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
picturePrompt = input("Download Pictures? (Y/N): ")

if picturePrompt == 'Y' or picturePrompt == 'y':
    downloadPictures(meterDate, 'North', pictureLoc)
    downloadPictures(meterDate, 'South', pictureLoc)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

executionTime = (time() - startTime)
print('Execution time in seconds: ' + str(executionTime))